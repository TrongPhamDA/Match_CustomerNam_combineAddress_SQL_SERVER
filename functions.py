import pandas as pd
from sqlalchemy import create_engine
import urllib
import sql_user.sql_user_login as login
import math
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
from unidecode import unidecode
import re

import os
from openpyxl import load_workbook

def distance_km(latA, longA, latB, longB):
    """
    Tính khoảng cách giữa hai điểm trên bề mặt trái đất theo đơn vị km.

    Args:
        latA (float): Vĩ độ của điểm A.
        longA (float): Kinh độ của điểm A.
        latB (float): Vĩ độ của điểm B.
        longB (float): Kinh độ của điểm B.

    Returns:
        float: Khoảng cách giữa hai điểm theo đơn vị km.
    """
    # Đổi độ từ độ sang radian
    latA_rad = math.radians(latA)
    longA_rad = math.radians(longA)
    latB_rad = math.radians(latB)
    longB_rad = math.radians(longB)
    
    # Tính sự khác biệt giữa các vĩ độ và kinh độ
    delta_lat = latB_rad - latA_rad
    delta_long = longB_rad - longA_rad
    
    # Công thức Haversine
    a = math.sin(delta_lat / 2) ** 2 + math.cos(latA_rad) * math.cos(latB_rad) * math.sin(delta_long / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    # Bán kính trái đất (trung bình) tính bằng km
    earth_radius_km = 6371.0
    
    # Tính khoảng cách
    distance = earth_radius_km * c
    
    return distance
    pass


def TFIDF_CosineSimilarity(input_customer_name: str, ref_customer_names: list) -> float:
    """
    Tính toán độ tương đồng cosine sử dụng TF-IDF giữa tên khách hàng đầu vào và một danh sách tên khách hàng tham khảo.

    Args:
        input_customer_name (str): Tên khách hàng đầu vào.
        ref_customer_names (list): Danh sách các tên khách hàng tham khảo.

    Returns:
        float: Độ tương đồng cosine giữa tên khách hàng đầu vào và các tên khách hàng tham khảo.
    """
    if not input_customer_name:
        return 0.0
    
    # Chuẩn bị dữ liệu cho TF-IDF
    corpus = [input_customer_name] + ref_customer_names
    
    # Khởi tạo vectorizer TF-IDF
    vectorizer = TfidfVectorizer()
    
    # Chuyển đổi dữ liệu thành các vector TF-IDF
    tfidf_matrix = vectorizer.fit_transform(corpus)
    
    # Tính toán độ tương đồng cosine
    cosine_sim = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:])
    
    # Trả về độ tương đồng cao nhất từ danh sách tham khảo
    return cosine_sim.max() if cosine_sim.size > 0 else 0.0


def unicode_convert_to_ascii(text):
    """
    Chuyển đổi các ký tự Unicode trong chuỗi thành ký tự ASCII không dấu
    và chuyển đổi toàn bộ chuỗi thành chữ thường. Xử lý giá trị NaN và không phải chuỗi.

    Args:
        text (str): Chuỗi chứa ký tự Unicode.

    Returns:
        str: Chuỗi đã được chuyển đổi thành ký tự ASCII và chữ thường.
    """
    if pd.isna(text):  # Kiểm tra nếu giá trị là NaN
        return ''
    if isinstance(text, str):  # Kiểm tra xem text có phải là chuỗi không
        return unidecode(text).lower()  # Chuyển đổi thành ASCII và chữ thường
    return str(text)  # Chuyển đổi giá trị không phải chuỗi thành chuỗi

import re

def remove_special_characters(text):
    """
    Loại bỏ tất cả các ký tự đặc biệt từ chuỗi, chỉ giữ lại chữ cái và số.
    
    Args:
        text (str): Chuỗi đầu vào chứa các ký tự đặc biệt.
    
    Returns:
        str: Chuỗi đã loại bỏ ký tự đặc biệt, chỉ còn chữ cái và số.
    """
    if pd.isna(text):  # Kiểm tra nếu giá trị là NaN
        return ''
    if isinstance(text, str):  # Kiểm tra xem text có phải là chuỗi không
        # Loại bỏ tất cả các ký tự không phải chữ cái và số
        return re.sub(r'[^a-zA-Z0-9\s]', '', text)
    return str(text)  # Chuyển đổi giá trị không phải chuỗi thành chuỗi


def remove_unwanted_words(text):
    """
    Loại bỏ các từ không cần thiết từ chuỗi đầu vào nếu các từ đó đứng riêng biệt.
    
    :param text: Chuỗi văn bản cần xử lý.
    :return: Chuỗi văn bản sau khi loại bỏ các từ không cần thiết.
    """
    # Danh sách các từ cần bỏ
    words_to_remove = [
        'tinh', 'thanh pho', 'thi xa', 'thi tran', 'huyen', 'quan',
        'phuong', 'xa', 'thon', 'ap', 'duong', 'so', 'tinh lo',
        'quoc lo', 'so nha', 'duong so', 'tt', 'tx', 'tp', 'h', 'p', 'q', 'x'
    ]
    
    # Tạo biểu thức chính quy cho các từ cần bỏ, đảm bảo rằng từ phải đứng riêng biệt
    pattern = r'\b(?:' + '|'.join(map(re.escape, words_to_remove)) + r')\b'
    
    # Loại bỏ các từ không cần thiết
    cleaned_text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    
    # Xóa khoảng trắng thừa và trả về kết quả
    return ' '.join(cleaned_text.split())


def preprocess_text(text):
    """
    Chuyển đổi ký tự Unicode sang ASCII và loại bỏ ký tự đặc biệt.

    Args:
        text (str): Chuỗi đầu vào để xử lý.

    Returns:
        str: Chuỗi đã được xử lý.
    """
    return remove_unwanted_words(remove_special_characters(unicode_convert_to_ascii(text)))


def query_exportExcel(sql_file_path: str, output_file: str, sheet_name: str):
    """
    Thực hiện truy vấn SQL từ file và ghi kết quả vào file Excel.

    Args:
        sql_file_path (str): Đường dẫn đến file SQL chứa truy vấn.
        output_file (str): Đường dẫn đến file Excel để lưu kết quả.
        sheet_name (str): Tên sheet trong file Excel để lưu dữ liệu.
    """
    # Đọc nội dung truy vấn SQL từ file
    with open(sql_file_path, 'r', encoding='utf-8-sig') as file:
        query_input = file.read()
    
    print("Truy vấn SQL:")
    print(query_input)

    # Thiết lập chuỗi kết nối sử dụng SQLAlchemy
    connection_string = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={login.server};"
        f"DATABASE={login.database};"
        f"UID={login.username};"
        f"PWD={login.password};"
    )
    connection_uri = f"mssql+pyodbc:///?odbc_connect={urllib.parse.quote_plus(connection_string)}"

    # Tạo engine sử dụng SQLAlchemy
    engine = create_engine(connection_uri)

    # Kết nối và thực hiện truy vấn
    try:
        with engine.connect() as connection:
            # Thực hiện truy vấn và đưa kết quả vào DataFrame
            df = pd.read_sql(query_input, connection)
            
            # Ghi kết quả vào file Excel
            df.to_excel(output_file, sheet_name=sheet_name, index=False)
            print(f"Kết quả đã được ghi vào {output_file}")
            
    finally:
        engine.dispose()
        print("Đã đóng kết nối.")

def levenshtein_distance(s1, s2):
    """Tính khoảng cách Levenshtein giữa hai chuỗi s1 và s2."""
    len_s1, len_s2 = len(s1), len(s2)
    
    # Tạo bảng khoảng cách
    dp = [[0] * (len_s2 + 1) for _ in range(len_s1 + 1)]
    
    # Khởi tạo bảng
    for i in range(len_s1 + 1):
        dp[i][0] = i
    for j in range(len_s2 + 1):
        dp[0][j] = j
    
    # Tính toán khoảng cách Levenshtein
    for i in range(1, len_s1 + 1):
        for j in range(1, len_s2 + 1):
            cost = 0 if s1[i - 1] == s2[j - 1] else 1
            dp[i][j] = min(dp[i - 1][j] + 1,      # Xóa
                           dp[i][j - 1] + 1,      # Thêm
                           dp[i - 1][j - 1] + cost)  # Thay thế
    
    return dp[len_s1][len_s2]


def write_toExcel_logfile(output_log_file, input_row, df_ref_temp, closest_distance_km):
    """
    Ghi log vào file Excel. Nếu file đã tồn tại thì ghi tiếp, nếu chưa thì tạo file mới.
    Tên sheet được giữ nguyên là "output_log".
    
    :param output_log_file: Đường dẫn đến file log.
    :param input_row: Dữ liệu từ input cần ghi log.
    :param df_ref_temp: Dữ liệu từ DataFrame reference cần ghi log (có thể chứa nhiều records).
    :param closest_distance_km: Khoảng cách gần nhất cần ghi log.
    """
    # Nếu file chưa tồn tại, tạo file mới với các cột cụ thể
    if not os.path.exists(output_log_file):
        # Các cột cần ghi
        columns = [
            'ROW_NO', 'CUSTOMER_NAME', 'CUSTOMER_ADDRESS', 'CUSTOMER_LAT', 'CUSTOMER_LONG', 'CUSTOMER_CODE', 
            'REF_CUSTOMER_NAME', 'REF_CUSTOMER_ADDRESS', 'MATCH_DISTANCE_KM', 'SEARCH_NAME_FIXED', 
            'SEARCH_ADDRESS_FIXED', 'SEARCH_COMBINE_FIXED', 'MATCH_NAME_SCORE', 'MATCH_NAME_SCORE_2',
            'MATCH_ADDRESS_SCORE', 'MATCH_COMBINE_SCORE', 'MATCH_FINAL_SCORE', 'CLOSEST_DISTANCE_KM'
        ]
        
        # Tạo file Excel với DataFrame trống có các cột và sheet name là "output_log"
        pd.DataFrame(columns=columns).to_excel(output_log_file, index=False, sheet_name='output_log')

    # Chuẩn bị dữ liệu để ghi
    records = []
    for _, row in df_ref_temp.iterrows():
        record = {
            'ROW_NO': input_row['ROW_NO'],
            'CUSTOMER_NAME': input_row['CUSTOMER_NAME'],
            'CUSTOMER_ADDRESS': input_row['CUSTOMER_ADDRESS'],
            'CUSTOMER_LAT': input_row['CUSTOMER_LAT'],
            'CUSTOMER_LONG': input_row['CUSTOMER_LONG'],
            'CUSTOMER_CODE': row['CUSTOMER_CODE'],
            'REF_CUSTOMER_NAME': row['CUSTOMER_NAME'],
            'REF_CUSTOMER_ADDRESS': row['CUSTOMER_ADDRESS'],
            'MATCH_DISTANCE_KM': row['MATCH_DISTANCE_KM'],
            'SEARCH_NAME_FIXED': row['SEARCH_NAME_FIXED'],
            'SEARCH_ADDRESS_FIXED': row['SEARCH_ADDRESS_FIXED'],
            'SEARCH_COMBINE_FIXED': row['SEARCH_COMBINE_FIXED'],
            'MATCH_NAME_SCORE': row['MATCH_NAME_SCORE'],
            'MATCH_NAME_SCORE_2': row['MATCH_NAME_SCORE_2'],
            'MATCH_ADDRESS_SCORE': row['MATCH_ADDRESS_SCORE'],
            'MATCH_COMBINE_SCORE': row['MATCH_COMBINE_SCORE'],
            'MATCH_FINAL_SCORE': row['MATCH_FINAL_SCORE'],
            'CLOSEST_DISTANCE_KM': closest_distance_km
        }
        records.append(record)
    
    # Convert records thành DataFrame để append vào file log
    log_df = pd.DataFrame(records)
    
    # Sử dụng openpyxl để mở file và append dữ liệu
    if os.path.exists(output_log_file):
        book = load_workbook(output_log_file)
        writer = pd.ExcelWriter(output_log_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book
        sheet = book['output_log']
        startrow = sheet.max_row
        log_df.to_excel(writer, sheet_name='output_log', index=False, header=False, startrow=startrow)
    else:
        log_df.to_excel(output_log_file, sheet_name='output_log', index=False)
    
    # Lưu và đóng writer
    writer.save()
    writer.close()